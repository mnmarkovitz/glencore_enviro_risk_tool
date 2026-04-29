"""
Export a comprehensive companion Excel workbook — "Quick Reference" for
Responsible Sourcing analysts who want the scored output without running
the full Streamlit tool.

Output: Quick_Reference.xlsx (in project root)

Sheets:
  1.  README                          - what this is + bucket legend + caveats
  2.  User Guide                      - end-user walkthrough
  3.  Country × Risk Heatmap          - color-coded matrix + bar chart
  4.  Full Ranked Results             - every scored row, filterable, color-coded
  5.  Risk Library                    - 15 risks with definition, KPIs, processes
  6.  Risk × Process Matrix           - intensity (1–5) of each process per risk
  7.  Commodity Producers             - USGS top producers + critical-mineral flag + chart
  8.  Country Indicators              - all raw indicators per country, sourced
  9.  Soil Vulnerability (SoilGrids)  - pH, SOC, CEC, derived vulnerability
  10. Water Stress (Aqueduct)         - WRI Aqueduct 4.0 country scores
  11. Glencore-Owned Assets           - public industrial assets from annual report
  12. Noise Baseline                  - NIOSH dBA per mining activity
  13. Methodology + Scoring Weights   - formulas, normalization, weights (combined)
  14. Supplier Engagement Tiers       - OSDR → SAQ → onsite → OGA → CAP
  15. Data Sources                    - all hyperlinked citations

Run whenever the scoring inputs change:
    python scripts/08_export_quick_reference.py
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "app"))
from scoring import compute, _load  # noqa: E402

OUTPUT = ROOT / "Quick_Reference.xlsx"

# Bucket colors — match the Streamlit app exactly
BUCKET_FILLS = {
    "Low":      PatternFill("solid", start_color="FF4CAF50", end_color="FF4CAF50"),
    "Moderate": PatternFill("solid", start_color="FFFFC107", end_color="FFFFC107"),
    "High":     PatternFill("solid", start_color="FFFF9800", end_color="FFFF9800"),
    "Critical": PatternFill("solid", start_color="FFE53935", end_color="FFE53935"),
    "No data":  PatternFill("solid", start_color="FFBDBDBD", end_color="FFBDBDBD"),
}
HEADER_FILL = PatternFill("solid", start_color="FF00A9A5", end_color="FF00A9A5")  # Glencore teal
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
BUCKET_FONT_LIGHT = Font(color="FFFFFFFF", bold=True)
BUCKET_FONT_DARK = Font(color="FF000000", bold=True)
THIN_BORDER = Border(*(Side(border_style="thin", color="FFCCCCCC"),) * 4)


def _bucket_of(overall):
    if pd.isna(overall): return "No data"
    if overall <= 4: return "Low"
    if overall <= 9: return "Moderate"
    if overall <= 14: return "High"
    return "Critical"


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_source_note(ws, text, n_cols, hyperlinks=None):
    """Insert a sources note as the first row of a sheet (above the header)."""
    ws.insert_rows(1)
    ws["A1"] = text
    ws["A1"].font = Font(italic=True, color="FF555555", size=10)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{end_col}1")
    ws.row_dimensions[1].height = 36
    if hyperlinks:
        for url, label in hyperlinks:
            pass  # markdown-style links don't render natively; we keep URLs in source_note text


# Mapping of country-indicator field names to (Display, Source label, URL)
INDICATOR_META = {
    "iso3":                          ("ISO-3 code",                   "ISO 3166-1 alpha-3", ""),
    "country":                       ("Country",                       "", ""),
    "epi_overall_2024":              ("Yale EPI 2024 overall (0–100, higher better)",
                                       "Yale Center for Environmental Law & Policy",
                                       "https://epi.yale.edu"),
    "epi_ecosystem_vitality":        ("Yale EPI 2024 — Ecosystem Vitality sub-index (0–100)",
                                       "Yale EPI 2024", "https://epi.yale.edu"),
    "epi_biodiversity_habitat":      ("Yale EPI 2024 — Biodiversity & Habitat sub-index (0–100)",
                                       "Yale EPI 2024", "https://epi.yale.edu"),
    "epi_air_quality":               ("Yale EPI 2024 — Air Quality sub-index (0–100)",
                                       "Yale EPI 2024", "https://epi.yale.edu"),
    "epi_waste_management":          ("Yale EPI 2024 — Waste Management sub-index (0–100)",
                                       "Yale EPI 2024", "https://epi.yale.edu"),
    "epi_heavy_metals":              ("Yale EPI 2024 — Heavy Metals sub-index (0–100)",
                                       "Yale EPI 2024", "https://epi.yale.edu"),
    "who_pm25_annual_ugm3":          ("WHO Ambient PM2.5 (annual mean, µg/m³)",
                                       "WHO Ambient Air Quality Database",
                                       "https://www.who.int/data/gho/data/themes/air-pollution/who-air-quality-database"),
    "wb_co2_t_per_capita":           ("World Bank — CO2 emissions (metric tonnes per capita)",
                                       "World Bank Open Data (EN.ATM.CO2E.PC)",
                                       "https://data.worldbank.org/indicator/EN.ATM.CO2E.PC"),
    "gfw_tree_cover_loss_pct_2023":  ("Global Forest Watch — tree cover loss (% of country area, 2023)",
                                       "Global Forest Watch / WRI",
                                       "https://www.globalforestwatch.org"),
    "iucn_threatened_species":       ("IUCN Red List — threatened species count in country (CR + EN + VU)",
                                       "IUCN Red List API", "https://apiv3.iucnredlist.org"),
    "wdpa_protected_pct":            ("UNEP-WCMC WDPA — % terrestrial area protected",
                                       "Protected Planet (WDPA)",
                                       "https://www.protectedplanet.net"),
    "tsf_count":                     ("Global Tailings Portal — number of tailings storage facilities",
                                       "GRID-Arendal Global Tailings Portal",
                                       "https://tailing.grida.no"),
    "tsf_max_very_high_or_extreme":  ("GTP — count of TSFs with consequence class Very High / Extreme",
                                       "GRID-Arendal Global Tailings Portal",
                                       "https://tailing.grida.no"),
    "wb_wgi_gov_effectiveness":      ("World Bank WGI — Government Effectiveness (-2.5 to +2.5)",
                                       "World Bank Worldwide Governance Indicators",
                                       "https://www.worldbank.org/en/publication/worldwide-governance-indicators"),
    "wb_wgi_regulatory_quality":     ("World Bank WGI — Regulatory Quality (-2.5 to +2.5)",
                                       "World Bank Worldwide Governance Indicators",
                                       "https://www.worldbank.org/en/publication/worldwide-governance-indicators"),
    "unesco_heritage_sites":         ("UNESCO World Heritage — total inscribed sites in country",
                                       "UNESCO World Heritage Centre",
                                       "https://whc.unesco.org"),
    "unesco_heritage_in_danger":     ("UNESCO World Heritage — sites currently listed 'in danger'",
                                       "UNESCO World Heritage Centre",
                                       "https://whc.unesco.org"),
    "inform_risk_2024":              ("INFORM Risk Index 2024 (0–10)",
                                       "EC Joint Research Centre INFORM Risk",
                                       "https://drmkc.jrc.ec.europa.eu/inform-index"),
    "basel_hazwaste_kt_per_yr":      ("UNEP Basel Convention — national hazardous waste generation (kt/yr)",
                                       "Basel Convention", "http://www.basel.int"),
    "source_note":                   ("Source note (for this row)", "", ""),
    "cahra_flag":                    ("Glencore CAHRA flag (Y/N)",
                                       "Glencore CAHRA List 2025", "https://www.glencore.com"),
    "cahra_regions":                 ("Glencore CAHRA regions (sub-national)",
                                       "Glencore CAHRA List 2025", "https://www.glencore.com"),
    "nrgi_rgi_score_0_100":          ("NRGI Resource Governance Index 2021 (0–100, higher = better governance)",
                                       "Natural Resource Governance Institute",
                                       "https://resourcegovernanceindex.org"),
    "ej_atlas_conflict_count":       ("EJ Atlas — count of documented environmental conflicts",
                                       "Environmental Justice Atlas (EJOLT)",
                                       "https://ejatlas.org"),
    "soil_ph_0_5cm":                 ("ISRIC SoilGrids — topsoil pH (0–5cm depth)",
                                       "ISRIC SoilGrids 2.0", "https://soilgrids.org"),
    "soil_soc_g_per_kg":             ("ISRIC SoilGrids — topsoil organic carbon (g/kg)",
                                       "ISRIC SoilGrids 2.0", "https://soilgrids.org"),
    "soil_cec_cmol_per_kg":          ("ISRIC SoilGrids — topsoil CEC (cmol/kg)",
                                       "ISRIC SoilGrids 2.0", "https://soilgrids.org"),
    "soil_vulnerability_1_5":        ("Soil vulnerability score (1–5; derived from pH + SOC + CEC)",
                                       "Derived; see Methodology sheet", ""),
}


# -------------------------- sheet builders --------------------------

def sheet_readme(wb, risks, countries, producers, today):
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Environmental Risk — Quick Reference"
    ws["A1"].font = Font(bold=True, size=18, color="FF00A9A5")
    ws["A2"] = f"Generated: {today}   |   For: Glencore Group Responsible Sourcing team"
    ws["A2"].font = Font(italic=True, color="FF555555")

    ws["A4"] = "What this workbook is"
    ws["A4"].font = Font(bold=True, size=13)
    ws["A5"] = (
        "A precomputed snapshot of environmental risk scores for every commodity × country × "
        "mining process combination in the Glencore Responsible Sourcing team's tool. Use this "
        "Excel file for quick lookups when you don't need the full interactive Streamlit tool."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[5].height = 60

    ws["A7"] = "How to read the scores"
    ws["A7"].font = Font(bold=True, size=13)
    ws["A8"] = "Overall = Likelihood × Severity, range 1–25. Cells are color-coded by bucket:"
    ws["A10"] = "Low";      ws["A10"].fill = BUCKET_FILLS["Low"];      ws["A10"].font = BUCKET_FONT_LIGHT
    ws["B10"] = "Overall 1–4"
    ws["A11"] = "Moderate"; ws["A11"].fill = BUCKET_FILLS["Moderate"]; ws["A11"].font = BUCKET_FONT_DARK
    ws["B11"] = "Overall 5–9"
    ws["A12"] = "High";     ws["A12"].fill = BUCKET_FILLS["High"];     ws["A12"].font = BUCKET_FONT_DARK
    ws["B12"] = "Overall 10–14"
    ws["A13"] = "Critical"; ws["A13"].fill = BUCKET_FILLS["Critical"]; ws["A13"].font = BUCKET_FONT_LIGHT
    ws["B13"] = "Overall 15–25"

    ws["A15"] = "Sheets"
    ws["A15"].font = Font(bold=True, size=13)
    ws["A16"] = "README"
    ws["B16"] = "This page."
    ws["A17"] = "Country × Risk"
    ws["B17"] = "Heatmap: max Overall risk score per country × each environmental risk. Best for scanning exposure."
    ws["A18"] = "Full Ranked Results"
    ws["B18"] = "Every scored combination with sources. Use Excel's filter to slice."
    ws["A19"] = "Data Sources"
    ws["B19"] = "Hyperlinked list of every public dataset the scores come from."

    ws["A21"] = "Methodology (brief)"
    ws["A21"].font = Font(bold=True, size=13)
    ws["A22"] = "Likelihood = 0.4 × Process Intrinsic Risk + 0.6 × Country Hazard Score (both 1–5)"
    ws["A23"] = "Severity   = 0.5 × Ecological Sensitivity + 0.5 × Regulatory Strictness (both 1–5)"
    ws["A24"] = "Overall    = Likelihood × Severity"
    ws["A25"] = (
        "Full methodology, formulas, and normalization rules are in the Streamlit tool's "
        "Methodology tab and in docs/METHODOLOGY.md."
    )
    ws["A25"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[25].height = 40

    ws["A27"] = "Important caveats"
    ws["A27"].font = Font(bold=True, size=13)
    ws["A28"] = (
        "This workbook is a static snapshot. The interactive Streamlit tool is the Source of Truth. "
        "For map visualization, supplier overlays, and drill-down, use the Streamlit tool — not this file.\n\n"
        "Cells showing '—' mean the underlying public dataset has NO value for that country and risk. "
        "Examples: Noise pollution has no global country-level dataset (hence all '—' in the Hazard "
        "columns for that risk), and UNESCO heritage-in-danger counts are zero for many countries. "
        "In those cases, Likelihood falls back to the Process Intrinsic Risk alone."
    )
    ws["A28"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[28].height = 140

    ws["A30"] = "Built for the Glencore Group Responsible Sourcing team"
    ws["A30"].font = Font(italic=True)
    ws["A31"] = "in collaboration with the NYU SPS Center for Global Affairs Consulting Practicum."
    ws["A31"].font = Font(italic=True, color="FF555555")

    _autosize(ws, [18, 80])


def sheet_country_risk(wb, df, risks):
    ws = wb.create_sheet("Country × Risk")
    risk_ids = risks["risk_id"].tolist()
    risk_labels = dict(zip(risks["risk_id"], risks["risk_type"]))

    # Build max Overall per country × risk (all commodities, all processes, applies==Y)
    df_y = df[df["applies"] == "Y"].copy()
    pivot = (
        df_y[df_y["risk_id"].isin(risk_ids)]
        .groupby(["country", "iso3", "cahra_flag", "risk_id"])["overall_1_25"]
        .max().unstack("risk_id")
    )
    pivot = pivot.reindex(columns=[rid for rid in risk_ids if rid in pivot.columns])
    # Sort countries by their worst Overall score descending
    pivot["_max"] = pivot.max(axis=1)
    pivot = pivot.sort_values("_max", ascending=False).drop(columns=["_max"])
    pivot = pivot.reset_index()

    # Header row
    headers = ["Country", "ISO", "CAHRA"] + [risk_labels[rid] for rid in pivot.columns[3:]]
    ws.append(headers)
    _style_header(ws[1])

    # Data rows
    for _, r in pivot.iterrows():
        row_data = [r["country"], r["iso3"], r["cahra_flag"]]
        row_data += [float(r[rid]) if pd.notna(r[rid]) else None for rid in pivot.columns[3:]]
        ws.append(row_data)
        row_idx = ws.max_row
        # CAHRA cell styling
        if r["cahra_flag"] == "Y":
            ws.cell(row=row_idx, column=3).fill = PatternFill("solid", start_color="FFFFD54F", end_color="FFFFD54F")
            ws.cell(row=row_idx, column=3).font = Font(bold=True)
        # Color each score cell
        for i in range(4, len(headers) + 1):
            val = ws.cell(row=row_idx, column=i).value
            if val is None: continue
            bucket = _bucket_of(val)
            ws.cell(row=row_idx, column=i).fill = BUCKET_FILLS[bucket]
            ws.cell(row=row_idx, column=i).font = BUCKET_FONT_LIGHT if bucket in ("Low", "Critical") else BUCKET_FONT_DARK
            ws.cell(row=row_idx, column=i).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=i).number_format = "0.0"

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, [32, 6, 8] + [16] * (len(headers) - 3))


def sheet_full_table(wb, df):
    ws = wb.create_sheet("Full Ranked Results")
    df_y = df[df["applies"] == "Y"].copy()
    cols = [
        "risk_type", "commodity", "country", "cahra_flag", "process",
        "country_hazard_raw", "country_hazard_norm_1_5",
        "likelihood_1_5", "severity_1_5", "overall_1_25", "risk_bucket",
        "process_intrinsic_1_5",
        "ecological_sensitivity_1_5", "regulatory_strictness_1_5",
        "likely_supplier_types", "country_hazard_source",
    ]
    df_y = df_y[cols].sort_values("overall_1_25", ascending=False)
    df_y.columns = [
        "Risk", "Commodity", "Country", "CAHRA", "Process",
        "Hazard Raw (source units)", "Hazard Normalized (1-5)",
        "Likelihood (1-5)", "Severity (1-5)", "Overall (1-25)", "Bucket",
        "Process Intrinsic (1-5)",
        "Eco Sensitivity (1-5)", "Regulatory Strict. (1-5)",
        "Likely Supplier Types", "Country Hazard Source",
    ]

    # Header
    ws.append(df_y.columns.tolist())
    _style_header(ws[1])

    # Data
    for _, row in df_y.iterrows():
        cells = [("—" if pd.isna(v) else v) for v in row.tolist()]
        ws.append(cells)
        r = ws.max_row
        if row["CAHRA"] == "Y":
            ws.cell(row=r, column=4).fill = PatternFill("solid", start_color="FFFFD54F", end_color="FFFFD54F")
            ws.cell(row=r, column=4).font = Font(bold=True)
        bucket = row["Bucket"]
        ws.cell(row=r, column=11).fill = BUCKET_FILLS.get(bucket, PatternFill())
        ws.cell(row=r, column=11).font = BUCKET_FONT_LIGHT if bucket in ("Low", "Critical") else BUCKET_FONT_DARK
        # Number formats for numeric columns (skip cells that are the "—" placeholder)
        for c in [6, 7, 8, 9, 10, 12, 13, 14]:
            if isinstance(ws.cell(row=r, column=c).value, (int, float)):
                ws.cell(row=r, column=c).number_format = "0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, [40, 18, 32, 8, 14, 18, 18, 14, 14, 14, 10, 18, 16, 20, 50, 50])


def _generic_table(wb, sheet_name, df, widths=None, hyperlink_cols=None, freeze="A2"):
    """Generic helper to drop a DataFrame onto a sheet with header styling."""
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    _style_header(ws[1])
    for _, row in df.iterrows():
        ws.append([("" if pd.isna(v) else v) for v in row.tolist()])
        r = ws.max_row
        if hyperlink_cols:
            for c in hyperlink_cols:
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="FF0366D6", underline="single")
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    if widths:
        _autosize(ws, widths)
    return ws


def sheet_risk_library(wb, risks, matrix, risk_supplier):
    ws = wb.create_sheet("Risk Library")
    ws.append(["Risk", "Definition", "Key KPIs (for SAQ)",
                "Driving processes (intensity 1-5)", "Likely supplier types",
                "Likelihood dataset", "Severity dataset"])
    _style_header(ws[1])
    rs_map = dict(zip(risk_supplier["risk_id"], risk_supplier["supplier_types"])) \
        if "risk_id" in risk_supplier.columns else {}
    for _, r in risks.iterrows():
        proc_rows = matrix[(matrix["risk_id"] == r["risk_id"]) & (matrix["applies"] == "Y")] \
            .sort_values("intrinsic_intensity_1_5", ascending=False)
        proc_str = "; ".join(f"{p['process']}: {int(p['intrinsic_intensity_1_5'])}"
                              for _, p in proc_rows.iterrows())
        ws.append([
            r["risk_type"], r["definition"], r["key_kpis"],
            proc_str,
            rs_map.get(r["risk_id"], r.get("likely_supplier_types", "")),
            r["likelihood_dataset"], r["severity_dataset"],
        ])
        row_idx = ws.max_row
        for c in [2, 3, 4, 5]:
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        # Hyperlink the dataset names back to their public URLs
        for col_idx, url_col in [(6, "likelihood_url"), (7, "severity_url")]:
            cell = ws.cell(row=row_idx, column=col_idx)
            url = r.get(url_col)
            if isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.font = Font(color="FF0366D6", underline="single", size=10)
        ws.row_dimensions[row_idx].height = 90
    ws.freeze_panes = "A2"
    _autosize(ws, [32, 60, 60, 40, 40, 28, 28])


def sheet_risk_process_matrix(wb, risks, matrix):
    ws = wb.create_sheet("Risk × Process Matrix")
    pivot = matrix.pivot(index="risk_id", columns="process",
                          values="intrinsic_intensity_1_5").fillna(0)
    # Reorder columns so processes follow value-chain order
    order = ["Mining", "Refining", "Smelting", "Recycling", "Marketing"]
    pivot = pivot.reindex(columns=[c for c in order if c in pivot.columns])
    # Replace risk_id with risk_type
    label_map = dict(zip(risks["risk_id"], risks["risk_type"]))
    pivot.index = pivot.index.map(label_map)
    pivot = pivot.reset_index().rename(columns={"risk_id": "Risk"})
    ws.append(list(pivot.columns))
    _style_header(ws[1])
    for _, r in pivot.iterrows():
        ws.append(r.tolist())
        row_idx = ws.max_row
        for col_i, val in enumerate(r.tolist()[1:], start=2):
            cell = ws.cell(row=row_idx, column=col_i)
            try:
                v = int(float(val))
            except (TypeError, ValueError):
                v = 0
            # 1-5 intensity color (low→high)
            colors = {0: "FFEEEEEE", 1: "FFC8E6C9", 2: "FFFFF59D",
                      3: "FFFFCC80", 4: "FFFF8A65", 5: "FFE53935"}
            cell.fill = PatternFill("solid", start_color=colors.get(v, "FFEEEEEE"),
                                     end_color=colors.get(v, "FFEEEEEE"))
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True, color="FFFFFFFF" if v >= 5 else "FF000000")
    ws.freeze_panes = "B2"
    _autosize(ws, [40] + [14] * (len(pivot.columns) - 1))


def sheet_commodity_producers(wb, producers):
    df = producers.copy()
    df.columns = ["Commodity", "Country", "ISO-3", "Rank", "Share % global",
                   "Source", "Critical Mineral", "Critical Source"]
    ws = _generic_table(wb, "Commodity Producers", df,
                         widths=[18, 30, 8, 8, 12, 28, 12, 60],
                         hyperlink_cols=[8])
    # Source row at top
    ws.insert_rows(1)
    ws["A1"] = ("Sources: USGS Mineral Commodity Summaries 2024 "
                 "(https://pubs.usgs.gov/periodicals/mcs2024) for metals; "
                 "BP Statistical Review of World Energy 2024 (https://www.energyinst.org/statistical-review) "
                 "+ IEA (https://www.iea.org) for coal and oil/gas. "
                 "Critical-mineral flag from USGS 2022 Critical Minerals List "
                 "(https://www.usgs.gov/news/national-news-release/us-geological-survey-releases-2022-list-critical-minerals).")
    ws["A1"].font = Font(italic=True, color="FF005F73", size=11)
    ws["A1"].hyperlink = "https://pubs.usgs.gov/periodicals/mcs2024"
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 60


def sheet_cahra_list(wb, countries):
    ws = wb.create_sheet("CAHRA Country List")
    cahra = countries[countries["cahra_flag"] == "Y"][["iso3", "country", "cahra_regions"]] \
        .sort_values("country")
    cahra.columns = ["ISO-3", "Country", "CAHRA regions"]
    ws.append(list(cahra.columns))
    _style_header(ws[1])
    for _, r in cahra.iterrows():
        ws.append(r.tolist())
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=2).fill = PatternFill(
            "solid", start_color="FFFFD54F", end_color="FFFFD54F")
        ws.cell(row=row_idx, column=2).font = Font(bold=True)
    ws.freeze_panes = "A2"
    _autosize(ws, [8, 32, 60])
    # Note row at top
    ws.insert_rows(1)
    ws["A1"] = (f"Source: Glencore CAHRA List 2025 (updated 27.02.2025) — "
                 f"{len(cahra)} countries flagged as Conflict-Affected & High-Risk Areas")
    ws["A1"].font = Font(italic=True, color="FF555555", size=11)
    ws.merge_cells("A1:C1")


def sheet_country_indicators(wb, countries):
    df = countries.copy().sort_values("country").reset_index(drop=True)
    ws = wb.create_sheet("Country Indicators")

    # Row 1: source note
    ws["A1"] = ("Per-column source: hover any header for the dataset name. Full URL list "
                "in the 'Data Sources' sheet. All values public; refresh via "
                "scripts/02_fetch_external_data.py.")
    ws["A1"].font = Font(italic=True, color="FF555555", size=10)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    end_col = get_column_letter(len(df.columns))
    ws.merge_cells(f"A1:{end_col}1")
    ws.row_dimensions[1].height = 36

    # Row 2: friendly header + comment with source URL
    ws.append([INDICATOR_META.get(c, (c, "", ""))[0] for c in df.columns])
    _style_header(ws[2])
    from openpyxl.comments import Comment
    for col_i, col_name in enumerate(df.columns, start=1):
        meta = INDICATOR_META.get(col_name)
        if meta and meta[1]:
            cell = ws.cell(row=2, column=col_i)
            cell.comment = Comment(f"{meta[1]}\n{meta[2]}", "Tool")

    # Data rows
    for _, row in df.iterrows():
        ws.append([("" if pd.isna(v) else v) for v in row.tolist()])
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{end_col}{ws.max_row}"
    widths = [6, 30] + [22] * (len(df.columns) - 2)
    _autosize(ws, widths)

    # Row 3 onward: highlight CAHRA rows (look up cahra_flag column)
    cahra_col = list(df.columns).index("cahra_flag") + 1 if "cahra_flag" in df.columns else None
    if cahra_col:
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=cahra_col).value == "Y":
                ws.cell(row=r, column=2).fill = PatternFill(
                    "solid", start_color="FFFFD54F", end_color="FFFFD54F")
                ws.cell(row=r, column=2).font = Font(bold=True)


def sheet_user_guide(wb):
    """Render USER_GUIDE.md as a sheet (markdown stripped to plain headings + bullets)."""
    ws = wb.create_sheet("User Guide")
    ws["A1"] = "Tool — User Guide"
    ws["A1"].font = Font(bold=True, size=18, color="FF00A9A5")
    ws["A2"] = ("A practical walkthrough for the Responsible Sourcing analyst. "
                 "Read this once and you can run the tool end-to-end.")
    ws["A2"].font = Font(italic=True, color="FF555555")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

    sections = [
        ("What this tool does (1 sentence)",
         "Given a commodity, country, and process stage, the tool returns a ranked list "
         "of environmental risks — each with a Likelihood × Severity score drawn from public "
         "datasets — plus a map, a heatmap, and everything you need to scope a Supplier "
         "Questionnaire (SAQ)."),
        ("When to use it",
         "Onboarding a new supplier · Annual SCDD review · Scoping an onsite visit · "
         "Drafting a CAP · Preparing for CSRD/CSDDD compliance reporting."),
        ("How to read a score",
         "All scores 1–5 (5 = worst). Likelihood = 0.4 × Process Intrinsic + 0.6 × Country "
         "Hazard. Severity = 0.5 × Ecological Sensitivity + 0.5 × Regulatory Strictness. "
         "Overall = L × S, range 1–25. Buckets: 1–4 Low · 5–9 Moderate · 10–14 High · 15–25 "
         "Critical. Strict regulators raise Severity (penalty exposure)."),
        ("Quickstart — the first 5 minutes",
         "1) Pick a commodity in the Streamlit tool's sidebar. 2) Look at the Risk Matrix "
         "(Likelihood × Severity) heatmap. 3) Read the Ranked Risks table — sort by Overall. "
         "4) Click the drill-down to see raw indicator + public-source URL. 5) Open the "
         "Risk Library tab to grab SAQ KPIs."),
        ("Common workflows (Excel-only)",
         "(a) Open 'Country × Risk Heatmap' for a one-page country exposure scan. "
         "(b) Open 'Full Ranked Results' and use Excel's filter for commodity = X, "
         "country = Y. (c) Open 'Risk Library' to copy SAQ KPIs into your questionnaire. "
         "(d) Open 'Glencore-Owned Assets' to map the company's footprint to commodity / "
         "country combinations of interest."),
        ("Filters and CAHRA",
         "🚩 CAHRA flag = country (or sub-national region) on Glencore's CAHRA List 2025. "
         "Any CAHRA + High/Critical row should escalate to Tier 2 SAQ at minimum. "
         "⭐ Critical Mineral flag = USGS 2022 Critical Minerals List."),
        ("Supplier Engagement Tiers — where this tool fits",
         "Tier 1 (OSDR): this tool. Tier 2 (SAQ + clarification): tool's KPI list scopes "
         "the SAQ. Tier 3 (Onsite, internal SME): tool's Critical rows define agenda. "
         "Tier 4 (OGA, independent): same. Tier 5 (CAP + monitoring): KPIs become "
         "milestones. See the 'Supplier Engagement Tiers' sheet for full alignment to "
         "Glencore's SCDD M&M procedure."),
        ("Where to get help",
         "Why did a country score X? → drill-down in the Streamlit app or the "
         "'Full Ranked Results' sheet. Can I trust this score? → 'Data Sources' sheet. "
         "What's the math? → 'Methodology + Scoring Weights' sheet. "
         "Updates and integration paths are described in the project's handover document."),
    ]
    r = 4
    for title, body in sections:
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=13, color="FF005F73")
        r += 1
        ws.cell(row=r, column=1, value=body)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 60
        r += 2
    _autosize(ws, [120])


def sheet_soilgrids(wb):
    path = ROOT / "data" / "processed" / "soilgrids_country.csv"
    if not path.exists(): return
    df = pd.read_csv(path).sort_values("country")
    df.columns = ["ISO-3", "Country", "Topsoil pH (0-5cm)",
                   "Soil organic carbon (g/kg)", "Cation exchange capacity (cmol/kg)",
                   "Soil vulnerability (1-5)", "Source"]
    ws = _generic_table(wb, "Soil Vulnerability (SoilGrids)", df,
                         widths=[6, 28, 16, 18, 22, 18, 60])
    # Color the vulnerability column 1-5
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=6).value
        if isinstance(v, (int, float)):
            colors = {1: "FFC8E6C9", 2: "FFDCEDC8", 3: "FFFFF59D",
                      4: "FFFFCC80", 5: "FFFF8A65"}
            band = max(1, min(5, int(round(v))))
            ws.cell(row=r, column=6).fill = PatternFill(
                "solid", start_color=colors[band], end_color=colors[band])
            ws.cell(row=r, column=6).number_format = "0.00"
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
    # Note
    ws.insert_rows(1)
    ws["A1"] = ("Source: ISRIC SoilGrids 2.0 — https://soilgrids.org. "
                 "Vulnerability = mean of pH-distance, SOC-binding, CEC-buffering scores. "
                 "Higher = more mobile heavy metals.")
    ws["A1"].font = Font(italic=True, color="FF555555", size=11)
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 30


def sheet_aqueduct(wb):
    path = ROOT / "data" / "processed" / "aqueduct_country_scores.csv"
    if not path.exists(): return
    df = pd.read_csv(path).sort_values("name_0")
    df.columns = ["ISO-3", "Country", "UN region", "WB region",
                   "BWS — Baseline Water Stress (0-4)",
                   "DRR — Drought Risk (0-4)",
                   "RFR — Riverine Flood Risk (0-4)"]
    df = df.replace(-9999, "")
    ws = _generic_table(wb, "Water Stress (Aqueduct)", df,
                         widths=[6, 28, 14, 24, 22, 18, 22])
    # Source row at top
    ws.insert_rows(1)
    ws["A1"] = ("Source: WRI Aqueduct 4.0 — Country Rankings download (July 2023). "
                 "https://www.wri.org/applications/aqueduct/country-rankings/  ·  "
                 "License: CC BY 4.0  ·  Categories 0-4 (4 = extremely high). "
                 "Empty cell = no data published for that country.")
    ws["A1"].font = Font(italic=True, color="FF005F73", size=11)
    ws["A1"].hyperlink = "https://www.wri.org/applications/aqueduct/country-rankings/"
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 38


def sheet_glencore_assets(wb):
    path = ROOT / "data" / "processed" / "glencore_assets.csv"
    if not path.exists(): return
    df = pd.read_csv(path)
    df.columns = ["Asset", "Type", "Commodity", "Country", "Region / sub-national",
                   "ISO-3", "Lat", "Lon", "Ownership %", "Status",
                   "Commissioned / Acquired", "Source URL"]
    ws = _generic_table(wb, "Glencore-Owned Assets", df,
                         widths=[34, 28, 22, 22, 26, 6, 8, 8, 12, 26, 22, 50],
                         hyperlink_cols=[12])
    # Source row above
    ws.insert_rows(1)
    ws["A1"] = ("Source: Glencore corporate website (glencore.com), regional operations "
                 "pages, 2023 Annual Report, and country subsidiary sites (glencore.com.au, "
                 "glencore.ca, astronenergy.co.za). Always verify against the latest "
                 "annual report when updating.")
    ws["A1"].font = Font(italic=True, color="FF555555", size=10)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 36


def sheet_supplier_types(wb):
    path = ROOT / "data" / "processed" / "supplier_types.csv"
    if not path.exists(): return
    df = pd.read_csv(path)
    df.columns = ["Supplier type (high-risk category)",
                   "Possible environmental effect",
                   "Possible human-rights effect"]
    ws = _generic_table(wb, "Supplier Types", df, widths=[40, 60, 60])
    # Wrap text
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60


def sheet_risk_supplier(wb, risks, risk_supplier):
    if "risk_id" not in risk_supplier.columns: return
    df = risk_supplier.merge(risks[["risk_id", "risk_type"]], on="risk_id", how="left")
    df = df[["risk_type", "supplier_types"]]
    df.columns = ["Risk", "Likely supplier types"]
    ws = _generic_table(wb, "Risk → Supplier Types", df, widths=[40, 80])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 36


def sheet_noise(wb, noise):
    df = noise.copy()
    df.columns = ["Process", "Activity", "Typical dBA min", "Typical dBA max", "Source"]
    ws = _generic_table(wb, "Noise Baseline", df, widths=[14, 36, 16, 16, 60])
    # Color by max dBA
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=4).value
        if isinstance(v, (int, float)):
            color = "FFE53935" if v >= 105 else ("FFFF9800" if v >= 95 else
                     "FFFFC107" if v >= 85 else "FF4CAF50")
            ws.cell(row=r, column=4).fill = PatternFill("solid", start_color=color, end_color=color)
    ws.insert_rows(1)
    ws["A1"] = ("Source: NIOSH Mining Noise — https://www.cdc.gov/niosh/mining/topics/Noise.html"
                " + IFC EHS Guidelines (Base Metal Smelting & Refining). "
                "OSHA action level 85 dBA, PEL 90 dBA.")
    ws["A1"].font = Font(italic=True, color="FF555555", size=11)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 32


def sheet_methodology_and_weights(wb, weights):
    ws = wb.create_sheet("Methodology + Scoring Weights")
    ws["A1"] = "Scoring methodology"
    ws["A1"].font = Font(bold=True, size=18, color="FF00A9A5")
    rows = [
        ("Likelihood", "0.4 × Process Intrinsic Risk  +  0.6 × Country Hazard Score    →   1–5"),
        ("Severity",   "0.5 × Ecological Sensitivity   +  0.5 × Regulatory Strictness    →   1–5"),
        ("Overall",    "Likelihood × Severity     →   1–25"),
        ("",            ""),
        ("Buckets",    "1–4 Low · 5–9 Moderate · 10–14 High · 15–25 Critical"),
    ]
    r = 3
    for label, formula in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, color="FF005F73")
        ws.cell(row=r, column=2, value=formula)
        r += 1
    r += 2
    ws.cell(row=r, column=1, value="Why these weights?").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value=(
        "Process type determines whether a risk is possible at all (e.g., tailings only happen at "
        "mines), but the country's regulatory + ecological context determines whether a capable "
        "process actually causes harm. Two copper mines — Chile vs Zambia — have the same Process "
        "Intrinsic Risk for water depletion (5) but very different realized risk because Chile's "
        "Atacama is extremely water-stressed (Aqueduct BWS = 5) while Zambia is moderate (BWS = 2). "
        "Country context therefore weighs more (0.6) than process (0.4). Stricter regulators raise "
        "Severity because the tool measures penalty exposure, not pure ecological damage."
    ))
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 110
    r += 3
    ws.cell(row=r, column=1, value="Scoring weights (editable in scoring_weights.csv)").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Parameter").fill = HEADER_FILL
    ws.cell(row=r, column=1).font = HEADER_FONT
    ws.cell(row=r, column=2, value="Value").fill = HEADER_FILL
    ws.cell(row=r, column=2).font = HEADER_FONT
    ws.cell(row=r, column=3, value="Description").fill = HEADER_FILL
    ws.cell(row=r, column=3).font = HEADER_FONT
    weights_path = ROOT / "data" / "processed" / "scoring_weights.csv"
    if weights_path.exists():
        wdf = pd.read_csv(weights_path)
        for _, w in wdf.iterrows():
            r += 1
            ws.cell(row=r, column=1, value=w["parameter"])
            ws.cell(row=r, column=2, value=float(w["value"]))
            ws.cell(row=r, column=3, value=w["description"])
            ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 30
    _autosize(ws, [32, 12, 80])


def sheet_supplier_tiers(wb):
    ws = wb.create_sheet("Supplier Engagement Tiers")
    ws["A1"] = "Supplier Engagement Tiers — Glencore SCDD M&M Procedure"
    ws["A1"].font = Font(bold=True, size=18, color="FF00A9A5")
    ws["A2"] = ("Glencore's SCDD M&M Procedure follows the OECD Due Diligence Guidance (3rd ed.) "
                 "five-step framework. The tool automates Tier 1 (OSDR — Open-Source Desktop Research) "
                 "and feeds Tiers 2–4 with targeted questions and evidence requirements.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40
    headers = ["Tier", "Name", "What it does", "SCDD step", "Inputs", "Outputs", "Escalation rule"]
    ws.append([])
    ws.append(headers)
    _style_header(ws[4])
    tiers = [
        ("Tier 1", "OSDR — Open-Source Desktop Research",
         "Automated by THIS tool. Risk ranking from public datasets per "
         "(commodity × country × process). CAHRA flag, supplier-type cues, KPI list.",
         "Step 2A — Supplier/product scoping for SCDD",
         "Aqueduct, Yale EPI, WHO AAQ, Global Tailings Portal, IUCN Red List, WDPA, GFW, "
         "World Bank WGI, EDGAR, NIOSH, ISRIC SoilGrids, GEM, CAHRA list, USGS MCS + CMA",
         "Likelihood × Severity per row + CAHRA flag + likely supplier types + KPI watchlist",
         "Overall ≥ 10 (High/Critical) OR CAHRA-flagged OR red-flag location → escalate to Tier 2"),
        ("Tier 2", "SCDD Questionnaire (SAQ) + extended OSDR + supplier engagement",
         "Supplier completes SAQ targeted at the risks the tool flagged. Adverse-news screening, "
         "beneficial-ownership check, third-party assurance review (RMI / Copper Mark / LBMA / "
         "ICMM / ResponsibleSteel).",
         "Step 2B — SAQ + Step 2C Risk assessment + Step 2D Supplier engagement",
         "Supplier-completed SAQ, management-system documents, public policies, third-party "
         "assurances, beneficial-ownership records",
         "Evidence of EMS, traceability documents, corrective-action history, certifications",
         "Gaps or inconsistencies → Tier 3 (Onsite). Unresolved → Tier 4 (OGA). Risks managed → "
         "Approve. Severe → BAC review."),
        ("Tier 3", "Onsite visit by internal commercial team or HSEC&HR SME",
         "GRST sends an internal team to verify SAQ claims firsthand: site walkthrough, "
         "documentation review, interviews with operations staff. NOT an audit; lighter-touch "
         "than OGA.",
         "Step 2E — Onsite visits (Section 3.1.5 of the SCDD M&M procedure)",
         "GRST checklist; commercial team or internal HSEC&HR SME on site; site-visit report",
         "Firsthand observation; verification of management-system claims; documented site report",
         "Risks mitigated → end SCDD. Findings need independent verification → Tier 4 (OGA). "
         "CAP needed → Tier 5."),
        ("Tier 4", "On-the-Ground Assessment (OGA) — independent third-party",
         "When risks need independent verification beyond an internal onsite visit. OGA scope "
         "jointly approved by GRST + relationship owner + assessor; supplier consent required. "
         "Conducted by Group Internal Audit & Assurance (GIAA) or external consultant.",
         "Step 2F & 2G — On-the-Ground Assessment (Section 3.1.6 of the SCDD M&M procedure)",
         "OGA scope/protocol; supplier consent; independent assessor; site-based evaluation per "
         "OECD DDG / RMI All Minerals Standard",
         "Independent assessor report; signed nonconformances; verified traceability evidence",
         "Risks confirmed and unmanaged → Tier 5 (CAP). Supplier rejects OGA without reasonable "
         "explanation → suspend / terminate (BAC override possible)."),
        ("Tier 5", "Corrective Action Plan (CAP) + ongoing monitoring",
         "Time-bound remediation plan jointly designed by GRST + business + supplier. Ongoing "
         "monitoring with reporting cadence. CAP miss = reject 3P.",
         "Step 3 — CAPs + monitoring (Section 3.1.7 of the SCDD M&M procedure)",
         "CAP design (milestones, evidence requirements, reporting cadence)",
         "CAP document; monitoring reports; escalation triggers; re-assessment date; risk ranking "
         "update",
         "CAP met → 3P approved. CAP missed → 3P rejected; added to Declined Party List (DPL) "
         "by Head of Sustainability."),
    ]
    tier_colors = {"Tier 1": "FF4CAF50", "Tier 2": "FF1976D2",
                    "Tier 3": "FFFF9800", "Tier 4": "FF9C27B0",
                    "Tier 5": "FFE53935"}
    for t in tiers:
        ws.append(list(t))
        r = ws.max_row
        ws.cell(row=r, column=1).fill = PatternFill("solid",
            start_color=tier_colors[t[0]], end_color=tier_colors[t[0]])
        ws.cell(row=r, column=1).font = Font(bold=True, color="FFFFFFFF")
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for c in range(2, 8):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 110
    _autosize(ws, [10, 28, 50, 36, 50, 50, 36])




def sheet_sources(wb, risks):
    ws = wb.create_sheet("Data Sources")
    ws.append(["Risk", "Likelihood dataset", "Indicator", "URL",
                "Severity dataset", "Indicator", "URL"])
    _style_header(ws[1])
    for _, r in risks.iterrows():
        ws.append([r["risk_type"],
                    r["likelihood_dataset"], r["likelihood_indicator"], r["likelihood_url"],
                    r["severity_dataset"], r["severity_indicator"], r["severity_url"]])
        row = ws.max_row
        for col in [4, 7]:
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="FF0366D6", underline="single")
    ws.freeze_panes = "A2"
    _autosize(ws, [40, 28, 44, 60, 28, 44, 60])

    # Append a "Cross-cutting governance datasets" block beneath the risk list
    blank = ws.max_row + 2
    ws.cell(row=blank, column=1, value="Cross-cutting governance datasets (feed Regulatory Strictness for every risk):"
            ).font = Font(bold=True, color="FF005F73", size=12)
    governance_rows = [
        ("World Bank Worldwide Governance Indicators (WGI)",
         "Regulatory Quality + Government Effectiveness (-2.5 to +2.5)",
         "https://www.worldbank.org/en/publication/worldwide-governance-indicators"),
        ("Yale Environmental Performance Index (EPI 2024)",
         "Overall environmental governance score (0-100)", "https://epi.yale.edu"),
        ("NRGI Resource Governance Index 2021",
         "Extractive-sector specific governance (0-100)", "https://resourcegovernanceindex.org"),
        ("Environmental Justice Atlas (EJ Atlas / EJOLT)",
         "Count of documented environmental conflicts per country",
         "https://ejatlas.org"),
        ("Glencore CAHRA List 2025",
         "Conflict-Affected & High-Risk Areas (sub-national where applicable)",
         "https://www.glencore.com"),
    ]
    for name, indicator, url in governance_rows:
        blank += 1
        ws.cell(row=blank, column=1, value=name).font = Font(bold=True)
        ws.cell(row=blank, column=2, value=indicator)
        cell = ws.cell(row=blank, column=3, value=url)
        cell.hyperlink = url
        cell.font = Font(color="FF0366D6", underline="single")


def add_charts(wb, df, producers):
    """Insert Plotly-style native Excel charts on the heatmap + producers sheets."""
    from openpyxl.chart import BarChart, Reference, BarChart3D, PieChart

    # Top 10 countries by max Overall — bar chart on Country × Risk sheet
    if "Country × Risk" in wb.sheetnames:
        ws = wb["Country × Risk"]
        df_y = df[df["applies"] == "Y"].copy()
        top = (df_y.groupby("country")["overall_1_25"].max()
                   .sort_values(ascending=False).head(10).reset_index())
        # Append the chart-data block to the right of the table
        start_col = ws.max_column + 2
        ws.cell(row=1, column=start_col, value="Top 10 by max Overall")
        ws.cell(row=1, column=start_col).font = Font(bold=True, color="FF005F73")
        ws.cell(row=2, column=start_col, value="Country")
        ws.cell(row=2, column=start_col + 1, value="Max Overall")
        for i, row in top.iterrows():
            ws.cell(row=3 + i, column=start_col, value=row["country"])
            ws.cell(row=3 + i, column=start_col + 1, value=float(row["overall_1_25"]))
        chart = BarChart()
        chart.type = "bar"
        chart.style = 11
        chart.title = "Top 10 countries — max Overall risk"
        chart.x_axis.title = "Country"
        chart.y_axis.title = "Max Overall (1–25)"
        data = Reference(ws, min_col=start_col + 1, min_row=2, max_row=12, max_col=start_col + 1)
        cats = Reference(ws, min_col=start_col, min_row=3, max_row=12)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12; chart.width = 22
        ws.add_chart(chart, ws.cell(row=15, column=start_col).coordinate)

    # Production share by country for selected commodities — pie chart on Producers sheet
    if "Commodity Producers" in wb.sheetnames:
        ws = wb["Commodity Producers"]
        # Build a small Cobalt-only summary (good demo)
        cob = producers[producers["commodity"] == "Cobalt"][["country", "share_of_global_pct"]] \
              .sort_values("share_of_global_pct", ascending=False).head(10).reset_index(drop=True)
        if len(cob):
            start_col = ws.max_column + 2
            ws.cell(row=1, column=start_col, value="Cobalt — global production share by country")
            ws.cell(row=1, column=start_col).font = Font(bold=True, color="FF005F73")
            ws.cell(row=2, column=start_col, value="Country")
            ws.cell(row=2, column=start_col + 1, value="Share %")
            for i, r in cob.iterrows():
                ws.cell(row=3 + i, column=start_col, value=r["country"])
                ws.cell(row=3 + i, column=start_col + 1, value=float(r["share_of_global_pct"]))
            pie = PieChart()
            pie.title = "Cobalt — global production share (USGS MCS 2024)"
            data = Reference(ws, min_col=start_col + 1, min_row=2, max_row=12)
            labels = Reference(ws, min_col=start_col, min_row=3, max_row=12)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 12; pie.width = 18
            ws.add_chart(pie, ws.cell(row=15, column=start_col).coordinate)


def main():
    risks, matrix, countries, producers, noise, weights, _st = _load()
    risk_supplier = pd.read_csv(ROOT / "data" / "processed" / "risk_supplier_types.csv")
    today = date.today().isoformat()
    print(f"Computing scores for export... (today: {today})")
    df = compute()
    print(f"  {len(df):,} rows scored.")

    wb = Workbook()
    sheet_readme(wb, risks, countries, producers, today)
    sheet_user_guide(wb)
    sheet_country_risk(wb, df, risks)
    sheet_full_table(wb, df)
    sheet_risk_library(wb, risks, matrix, risk_supplier)
    sheet_risk_process_matrix(wb, risks, matrix)
    sheet_commodity_producers(wb, producers)
    sheet_country_indicators(wb, countries)
    sheet_soilgrids(wb)
    sheet_aqueduct(wb)
    sheet_glencore_assets(wb)
    sheet_noise(wb, noise)
    sheet_methodology_and_weights(wb, weights)
    sheet_supplier_tiers(wb)
    sheet_sources(wb, risks)
    wb.save(OUTPUT)
    print(f"✓ Wrote {OUTPUT.name}  ({OUTPUT.stat().st_size / 1024:.0f} KB) — {len(wb.sheetnames)} sheets")
    print(f"  Location: {OUTPUT}")


if __name__ == "__main__":
    main()
